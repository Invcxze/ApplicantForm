from io import BytesIO
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class FormSubmissionExporter:
    def export_submission_to_excel(self, submission):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Submission {submission.id}"

        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.append(["Поле", "Значение"])
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        for i, fv in enumerate(submission.values.select_related("field"), start=2):
            field_label = fv.field.label
            value = fv.text_value or fv.choice_value or ""
            ws.append([field_label, value])
            ws[f"A{i}"].border = thin_border
            ws[f"B{i}"].border = thin_border

        for col in range(1, 3):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws[f"{column_letter}{row}"].value or "")) for row in range(1, ws.max_row + 1))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        ws["A1"].alignment = center_alignment
        ws["B1"].alignment = center_alignment

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_submissions_to_zip(self, submissions):
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            for sub in submissions:
                excel = self.export_submission_to_excel(sub)
                archive.writestr(f"submission_{sub.id}.xlsx", excel.getvalue())
        buffer.seek(0)
        return buffer